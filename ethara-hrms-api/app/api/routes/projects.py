"""Project Governance & Budget Management routes.

- Project Master CRUD + admin-configurable columns (field defs)
- Excel/CSV bulk import & export
- Two-stage budget approval (CTO/COO -> Leadership)
- Analytics dashboard + Leadership view
- SLA escalation sweep
- Public (tokenized) budget approve/reject link handler (dormant while email off)

Mirrors the conventions of app/api/routes/reimbursements.py: dict serialization,
`require_permissions` gating, and in-app notifications via the workflows service.
"""

from __future__ import annotations

import csv
import io
import re
from collections import defaultdict
from datetime import UTC, date, datetime
from typing import Annotated, Any

from fastapi import APIRouter, Depends, File, HTTPException, Query, UploadFile, status
from fastapi.responses import HTMLResponse, StreamingResponse
from sqlalchemy import func, select
from sqlalchemy.orm import Session, selectinload

from app.api.deps import get_current_user, require_permissions
from app.core.config import get_settings
from app.core.database import get_db
from app.core.permissions import Permission
from app.core.security import decode_token, verify_token_hash
from app.db.models import (
    DinnerRequest,
    Project,
    ProjectBudget,
    ProjectFieldDef,
    ProjectLead,
    ReimbursementRequest,
    User,
    generate_id,
)
from app.schemas.projects import (
    ApproverConfigWrite,
    BudgetCreate,
    BudgetDecision,
    BudgetUpdate,
    FieldDefCreate,
    FieldDefReorder,
    FieldDefUpdate,
    ProjectCreate,
    ProjectUpdate,
    SlaConfigWrite,
)
from app.services import project_governance as pg
from app.services.audit import log_audit

router = APIRouter(prefix="/projects", tags=["projects"])
public_router = APIRouter(prefix="/public/projects", tags=["projects-public"])

PROJECT_TYPES = {"technical", "generalist"}
RFP_STATUSES = {"rfp", "production", "delivered"}
DELIVERY_STATUSES = {"ongoing", "completed"}


# --------------------------------------------------------------------------- #
# Parsing helpers
# --------------------------------------------------------------------------- #
def _slugify(value: str) -> str:
    slug = re.sub(r"[^a-z0-9]+", "_", (value or "").strip().lower()).strip("_")
    return slug or "field"


def _clean(value: Any) -> str | None:
    if value is None:
        return None
    text = str(value).strip()
    return text or None


def _parse_float(value: Any) -> float | None:
    text = _clean(value)
    if text is None:
        return None
    text = text.replace(",", "").replace("₹", "").replace("$", "").strip()
    try:
        return float(text)
    except ValueError:
        return None


def _parse_int(value: Any) -> int | None:
    f = _parse_float(value)
    return int(f) if f is not None else None


def _parse_date(value: Any) -> date | None:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    text = _clean(value)
    if not text:
        return None
    for fmt in ("%m/%d/%Y", "%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y", "%m-%d-%Y", "%d %b %Y"):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            continue
    return None


def _norm_enum(value: Any, allowed: set[str], default: str) -> str:
    text = (_clean(value) or "").lower()
    if text in allowed:
        return text
    # tolerate variants like "Completed / Ongoing"
    for token in allowed:
        if token in text:
            return token
    return default


# --------------------------------------------------------------------------- #
# Serialization
# --------------------------------------------------------------------------- #
def _serialize_lead(lead: ProjectLead) -> dict[str, Any]:
    return {
        "id": lead.id,
        "userId": lead.user_id,
        "name": lead.user.name if lead.user else None,
        "role": lead.role,
    }


def _serialize_project(db: Session, project: Project, *, with_spend: bool = True) -> dict[str, Any]:
    actual_spend = pg.compute_project_spend(db, project.id) if with_spend else (project.consumed_budget or 0.0)
    approved = project.approved_budget or 0.0
    leads = [_serialize_lead(lead) for lead in (project.leads or [])]
    latest_budget = max(
        (project.budgets or []), key=lambda b: (b.version or 0), default=None
    )
    return {
        "id": project.id,
        "internalName": project.internal_name,
        "externalName": project.external_name,
        "client": project.client,
        "platform": project.platform,
        "projectType": project.project_type,
        "rfpStatus": project.rfp_status,
        "deliveryStatus": project.delivery_status,
        "appsheetApproval": project.appsheet_approval,
        "trajectoryCostApproval": project.trajectory_cost_approval,
        "aht": project.aht,
        "targetVolume": project.target_volume,
        "deliveredVolume": project.delivered_volume,
        "dateOfDelivery": project.date_of_delivery.isoformat() if project.date_of_delivery else None,
        "tpmUserId": project.tpm_user_id,
        "tpmName": project.tpm.name if project.tpm else None,
        "fteDemand": project.fte_demand,
        "fteCount": project.fte_count,
        "internCount": project.intern_count,
        "totalMembers": project.total_members,
        "approvedBudget": approved,
        "consumedBudget": actual_spend,
        "remainingBudget": round(approved - actual_spend, 2),
        "currency": project.currency,
        "isArchived": project.is_archived,
        "customFields": project.custom_fields or {},
        "notes": project.notes,
        "leads": leads,
        "plNames": [lead["name"] for lead in leads if lead["role"] == "pl"],
        "latestBudgetStatus": latest_budget.status if latest_budget else None,
        "createdBy": project.created_by,
        "createdAt": project.created_at.isoformat() if project.created_at else None,
        "updatedAt": project.updated_at.isoformat() if project.updated_at else None,
    }


def _serialize_budget(budget: ProjectBudget) -> dict[str, Any]:
    actions = sorted(budget.actions or [], key=lambda a: a.created_at or datetime.min.replace(tzinfo=UTC))
    return {
        "id": budget.id,
        "projectId": budget.project_id,
        "version": budget.version,
        "amount": budget.amount,
        "currency": budget.currency,
        "period": budget.period,
        "justification": budget.justification,
        "status": budget.status,
        "proposedBy": budget.proposer.name if budget.proposer else None,
        "proposedById": budget.proposed_by,
        "submittedAt": budget.submitted_at.isoformat() if budget.submitted_at else None,
        "functionalApprover": budget.functional_approver.name if budget.functional_approver else None,
        "functionalApproverId": budget.functional_approver_id,
        "functionalDecision": budget.functional_decision,
        "functionalDecidedBy": budget.functional_decider.name if budget.functional_decider else None,
        "functionalDecidedAt": budget.functional_decided_at.isoformat() if budget.functional_decided_at else None,
        "functionalComment": budget.functional_comment,
        "leadershipDecision": budget.leadership_decision,
        "leadershipDecidedBy": budget.leadership_decider.name if budget.leadership_decider else None,
        "leadershipDecidedAt": budget.leadership_decided_at.isoformat() if budget.leadership_decided_at else None,
        "leadershipComment": budget.leadership_comment,
        "createdAt": budget.created_at.isoformat() if budget.created_at else None,
        "auditTrail": [
            {
                "id": a.id,
                "action": a.action,
                "stage": a.stage,
                "fromStatus": a.from_status,
                "toStatus": a.to_status,
                "comment": a.comment,
                "performedBy": a.performed_by_name,
                "performedByRole": a.performed_by_role,
                "createdAt": a.created_at.isoformat() if a.created_at else None,
            }
            for a in actions
        ],
    }


def _serialize_field_def(fd: ProjectFieldDef) -> dict[str, Any]:
    return {
        "id": fd.id,
        "key": fd.key,
        "label": fd.label,
        "dataType": fd.data_type,
        "options": fd.options or [],
        "group": fd.group,
        "orderIndex": fd.order_index,
        "isActive": fd.is_active,
    }


def _project_query():
    return select(Project).options(
        selectinload(Project.leads).selectinload(ProjectLead.user),
        selectinload(Project.budgets),
        selectinload(Project.tpm),
    )


def _load_project(db: Session, project_id: str) -> Project:
    project = db.scalar(_project_query().where(Project.id == project_id))
    if project is None:
        raise HTTPException(status_code=404, detail="Project not found.")
    return project


# --------------------------------------------------------------------------- #
# Lead handling
# --------------------------------------------------------------------------- #
def _replace_leads(db: Session, project: Project, *, tpm_user_id: str | None, lead_user_ids: list[str]) -> None:
    """Rebuild the project_leads rows from the TPM + PL ids supplied."""
    for existing in list(project.leads or []):
        db.delete(existing)
    db.flush()
    seen: set[tuple[str, str]] = set()
    if tpm_user_id:
        db.add(ProjectLead(id=generate_id(), project_id=project.id, user_id=tpm_user_id, role="tpm"))
        seen.add((tpm_user_id, "tpm"))
    for uid in lead_user_ids or []:
        if not uid or (uid, "pl") in seen:
            continue
        db.add(ProjectLead(id=generate_id(), project_id=project.id, user_id=uid, role="pl"))
        seen.add((uid, "pl"))


def _ensure_creator_lead(db: Session, project: Project, user: User) -> None:
    """Guarantee the creator can act on their own project even if not named TPM/PL."""
    if project.tpm_user_id == user.id:
        return
    if pg.is_project_lead(db, project.id, user.id):
        return
    db.add(ProjectLead(id=generate_id(), project_id=project.id, user_id=user.id, role="pl"))


# --------------------------------------------------------------------------- #
# Project Master CRUD
# --------------------------------------------------------------------------- #
@router.get("")
def list_projects(
    db: Annotated[Session, Depends(get_db)],
    current_user: Annotated[User, Depends(require_permissions(Permission.PROJECTS_READ))],
    include_archived: Annotated[bool, Query(alias="includeArchived")] = False,
) -> list[dict[str, Any]]:
    query = _project_query()
    if not include_archived:
        query = query.where(Project.is_archived.is_(False))
    query = query.order_by(Project.updated_at.desc())
    projects = list(db.scalars(query).unique())
    # Scope: Manager / PL-TPM see only projects they're tagged to; admin /
    # leadership / hr / office-admin see all.
    visible = pg.visible_project_ids(db, current_user)
    if visible is not None:
        projects = [p for p in projects if p.id in visible]
    return [_serialize_project(db, p) for p in projects]


@router.get("/options")
def project_options(
    db: Annotated[Session, Depends(get_db)],
    current_user: Annotated[User, Depends(get_current_user)],
) -> list[dict[str, Any]]:
    """Lightweight active-project list for expense-form dropdowns. Available to
    any authenticated user (exempt from module gating)."""
    rows = db.scalars(
        select(Project).where(Project.is_archived.is_(False)).order_by(Project.internal_name)
    )
    return [
        {"id": p.id, "internalName": p.internal_name, "externalName": p.external_name, "client": p.client}
        for p in rows
    ]


@router.post("", status_code=status.HTTP_201_CREATED)
def create_project(
    payload: ProjectCreate,
    db: Annotated[Session, Depends(get_db)],
    current_user: Annotated[User, Depends(require_permissions(Permission.PROJECTS_CREATE))],
) -> dict[str, Any]:
    project = Project(id=generate_id(), created_by=current_user.id)
    _apply_project_fields(db, project, payload, creating=True)
    db.add(project)
    db.flush()
    lead_ids = list(payload.lead_user_ids or [])
    if payload.leads:
        lead_ids.extend(lead.user_id for lead in payload.leads if lead.role != "tpm")
    _replace_leads(db, project, tpm_user_id=project.tpm_user_id, lead_user_ids=lead_ids)
    _ensure_creator_lead(db, project, current_user)
    db.flush()
    log_audit(
        db, entity_type="project", entity_id=project.id, action="project_created",
        actor=current_user, new_value={"internalName": project.internal_name},
    )
    db.commit()
    return _serialize_project(db, _load_project(db, project.id))


@router.patch("/{project_id}")
def update_project(
    project_id: str,
    payload: ProjectUpdate,
    db: Annotated[Session, Depends(get_db)],
    current_user: Annotated[User, Depends(require_permissions(Permission.PROJECTS_WRITE))],
) -> dict[str, Any]:
    project = _load_project(db, project_id)
    pg.assert_can_act_on_project(db, current_user, project)
    _apply_project_fields(db, project, payload, creating=False)
    # Only rebuild leads when the caller explicitly supplied them.
    if payload.lead_user_ids is not None or payload.leads is not None or payload.tpm_user_id is not None:
        lead_ids = list(payload.lead_user_ids or [])
        if payload.leads:
            lead_ids.extend(lead.user_id for lead in payload.leads if lead.role != "tpm")
        _replace_leads(db, project, tpm_user_id=project.tpm_user_id, lead_user_ids=lead_ids)
    db.add(project)
    db.flush()
    log_audit(
        db, entity_type="project", entity_id=project.id, action="project_updated",
        actor=current_user,
    )
    db.commit()
    return _serialize_project(db, _load_project(db, project.id))


@router.post("/{project_id}/archive")
def archive_project(
    project_id: str,
    db: Annotated[Session, Depends(get_db)],
    current_user: Annotated[User, Depends(require_permissions(Permission.PROJECTS_WRITE))],
    unarchive: Annotated[bool, Query()] = False,
) -> dict[str, Any]:
    project = _load_project(db, project_id)
    pg.assert_can_act_on_project(db, current_user, project)
    project.is_archived = not unarchive
    db.add(project)
    log_audit(
        db, entity_type="project", entity_id=project.id,
        action="project_unarchived" if unarchive else "project_archived", actor=current_user,
    )
    db.commit()
    return _serialize_project(db, _load_project(db, project.id))


@router.get("/{project_id}")
def get_project(
    project_id: str,
    db: Annotated[Session, Depends(get_db)],
    current_user: Annotated[User, Depends(require_permissions(Permission.PROJECTS_READ))],
) -> dict[str, Any]:
    project = _load_project(db, project_id)
    visible = pg.visible_project_ids(db, current_user)
    if visible is not None and project.id not in visible:
        raise HTTPException(status_code=404, detail="Project not found.")
    return _serialize_project(db, project)


def _apply_project_fields(db: Session, project: Project, payload: ProjectCreate | ProjectUpdate, *, creating: bool) -> None:
    data = payload.model_dump(exclude_unset=True, by_alias=False)
    simple = {
        "internal_name", "external_name", "client", "platform", "appsheet_approval",
        "trajectory_cost_approval", "aht", "target_volume", "delivered_volume",
        "tpm_user_id", "fte_demand", "fte_count", "intern_count", "total_members",
        "approved_budget", "consumed_budget", "notes",
    }
    for field in simple:
        if field in data:
            setattr(project, field, data[field])
    if "project_type" in data and data["project_type"]:
        project.project_type = _norm_enum(data["project_type"], PROJECT_TYPES, "technical")
    if "rfp_status" in data and data["rfp_status"]:
        project.rfp_status = _norm_enum(data["rfp_status"], RFP_STATUSES, "rfp")
    if "delivery_status" in data and data["delivery_status"]:
        project.delivery_status = _norm_enum(data["delivery_status"], DELIVERY_STATUSES, "ongoing")
    if "date_of_delivery" in data:
        project.date_of_delivery = _parse_date(data["date_of_delivery"])
    if "currency" in data and data["currency"]:
        project.currency = str(data["currency"]).upper()
    if "custom_fields" in data and data["custom_fields"] is not None:
        merged = dict(project.custom_fields or {})
        merged.update(data["custom_fields"])
        project.custom_fields = merged
    if creating and not project.total_members:
        project.total_members = (project.fte_count or 0) + (project.intern_count or 0) or None


# --------------------------------------------------------------------------- #
# Configurable columns (field defs)
# --------------------------------------------------------------------------- #
@router.get("/field-defs/all")
def list_field_defs(
    db: Annotated[Session, Depends(get_db)],
    current_user: Annotated[User, Depends(require_permissions(Permission.PROJECTS_READ))],
    include_inactive: Annotated[bool, Query(alias="includeInactive")] = False,
) -> list[dict[str, Any]]:
    query = select(ProjectFieldDef).order_by(ProjectFieldDef.order_index, ProjectFieldDef.label)
    if not include_inactive:
        query = query.where(ProjectFieldDef.is_active.is_(True))
    return [_serialize_field_def(fd) for fd in db.scalars(query)]


@router.post("/field-defs", status_code=status.HTTP_201_CREATED)
def create_field_def(
    payload: FieldDefCreate,
    db: Annotated[Session, Depends(get_db)],
    current_user: Annotated[User, Depends(require_permissions(Permission.PROJECTS_ADMIN))],
) -> dict[str, Any]:
    key = _slugify(payload.key or payload.label)
    if db.scalar(select(ProjectFieldDef.id).where(ProjectFieldDef.key == key)):
        raise HTTPException(status_code=409, detail=f"A column with key '{key}' already exists.")
    max_order = db.scalar(select(func.coalesce(func.max(ProjectFieldDef.order_index), 0))) or 0
    fd = ProjectFieldDef(
        id=generate_id(),
        key=key,
        label=payload.label,
        data_type=payload.data_type,
        options=payload.options or [],
        group=payload.group,
        order_index=payload.order_index if payload.order_index is not None else max_order + 1,
        created_by=current_user.id,
    )
    db.add(fd)
    db.commit()
    return _serialize_field_def(fd)


@router.patch("/field-defs/{field_id}")
def update_field_def(
    field_id: str,
    payload: FieldDefUpdate,
    db: Annotated[Session, Depends(get_db)],
    current_user: Annotated[User, Depends(require_permissions(Permission.PROJECTS_ADMIN))],
) -> dict[str, Any]:
    fd = db.get(ProjectFieldDef, field_id)
    if fd is None:
        raise HTTPException(status_code=404, detail="Column not found.")
    data = payload.model_dump(exclude_unset=True, by_alias=False)
    for attr in ("label", "data_type", "options", "group", "order_index", "is_active"):
        if attr in data and data[attr] is not None:
            setattr(fd, attr, data[attr])
    db.add(fd)
    db.commit()
    return _serialize_field_def(fd)


@router.delete("/field-defs/{field_id}", status_code=status.HTTP_204_NO_CONTENT)
def delete_field_def(
    field_id: str,
    db: Annotated[Session, Depends(get_db)],
    current_user: Annotated[User, Depends(require_permissions(Permission.PROJECTS_ADMIN))],
) -> None:
    fd = db.get(ProjectFieldDef, field_id)
    if fd is None:
        return
    # Soft-deactivate so historical custom_fields values are preserved.
    fd.is_active = False
    db.add(fd)
    db.commit()


@router.post("/field-defs/reorder")
def reorder_field_defs(
    payload: FieldDefReorder,
    db: Annotated[Session, Depends(get_db)],
    current_user: Annotated[User, Depends(require_permissions(Permission.PROJECTS_ADMIN))],
) -> list[dict[str, Any]]:
    for index, field_id in enumerate(payload.ordered_ids):
        fd = db.get(ProjectFieldDef, field_id)
        if fd is not None:
            fd.order_index = index + 1
            db.add(fd)
    db.commit()
    return list_field_defs(db, current_user)


# --------------------------------------------------------------------------- #
# Settings (approvers + SLA + email flag)
# --------------------------------------------------------------------------- #
@router.get("/settings/config")
def get_project_settings(
    db: Annotated[Session, Depends(get_db)],
    current_user: Annotated[User, Depends(require_permissions(Permission.PROJECTS_READ))],
) -> dict[str, Any]:
    approvers = pg.get_approver_config(db)
    sla = pg.get_sla_config(db)
    return {
        "approvers": {"technicalUserId": approvers.get("technical"), "generalistUserId": approvers.get("generalist")},
        "sla": sla,
        "emailEnabled": pg.email_enabled(db),
    }


@router.put("/settings/approvers")
def set_project_approvers(
    payload: ApproverConfigWrite,
    db: Annotated[Session, Depends(get_db)],
    current_user: Annotated[User, Depends(require_permissions(Permission.PROJECTS_ADMIN))],
) -> dict[str, Any]:
    pg.set_setting_value(
        db, key=pg.APPROVERS_KEY,
        value={"technical": payload.technical_user_id, "generalist": payload.generalist_user_id},
        actor=current_user, description="Functional (CTO/COO) budget approvers by project type.",
    )
    db.commit()
    return get_project_settings(db, current_user)


@router.put("/settings/sla")
def set_project_sla(
    payload: SlaConfigWrite,
    db: Annotated[Session, Depends(get_db)],
    current_user: Annotated[User, Depends(require_permissions(Permission.PROJECTS_ADMIN))],
) -> dict[str, Any]:
    current = pg.get_sla_config(db)
    value = {
        "budgetApprovalSlaHours": payload.budget_approval_sla_hours or current["budgetApprovalSlaHours"],
        "expenseApprovalSlaHours": payload.expense_approval_sla_hours or current["expenseApprovalSlaHours"],
    }
    pg.set_setting_value(db, key=pg.SLA_KEY, value=value, actor=current_user, description="Project approval SLA hours.")
    db.commit()
    return get_project_settings(db, current_user)


# --------------------------------------------------------------------------- #
# Budgets — two-stage approval
# --------------------------------------------------------------------------- #
@router.get("/{project_id}/budgets")
def list_budgets(
    project_id: str,
    db: Annotated[Session, Depends(get_db)],
    current_user: Annotated[User, Depends(require_permissions(Permission.PROJECTS_READ))],
) -> list[dict[str, Any]]:
    rows = db.scalars(
        select(ProjectBudget)
        .options(
            selectinload(ProjectBudget.actions),
            selectinload(ProjectBudget.proposer),
            selectinload(ProjectBudget.functional_approver),
            selectinload(ProjectBudget.functional_decider),
            selectinload(ProjectBudget.leadership_decider),
        )
        .where(ProjectBudget.project_id == project_id)
        .order_by(ProjectBudget.version.desc())
    )
    return [_serialize_budget(b) for b in rows]


@router.post("/{project_id}/budgets", status_code=status.HTTP_201_CREATED)
def create_budget(
    project_id: str,
    payload: BudgetCreate,
    db: Annotated[Session, Depends(get_db)],
    current_user: Annotated[User, Depends(require_permissions(Permission.PROJECTS_BUDGET_PROPOSE))],
) -> dict[str, Any]:
    project = _load_project(db, project_id)
    pg.assert_can_act_on_project(db, current_user, project)
    budget = ProjectBudget(
        id=generate_id(),
        project_id=project.id,
        version=pg.next_budget_version(db, project.id),
        amount=payload.amount,
        currency=(payload.currency or project.currency or "INR").upper(),
        period=payload.period,
        justification=payload.justification,
        status=pg.STATUS_DRAFT,
        proposed_by=current_user.id,
    )
    db.add(budget)
    db.flush()
    pg.log_budget_action(
        db, budget=budget, actor=current_user, action="created", stage="proposal",
        from_status=None, to_status=pg.STATUS_DRAFT,
    )
    db.commit()
    return _serialize_budget(_load_budget(db, budget.id))


@router.patch("/budgets/{budget_id}")
def update_budget(
    budget_id: str,
    payload: BudgetUpdate,
    db: Annotated[Session, Depends(get_db)],
    current_user: Annotated[User, Depends(require_permissions(Permission.PROJECTS_BUDGET_PROPOSE))],
) -> dict[str, Any]:
    budget = _load_budget(db, budget_id)
    project = _load_project(db, budget.project_id)
    pg.assert_can_act_on_project(db, current_user, project)
    if budget.status not in {pg.STATUS_DRAFT, pg.STATUS_REJECTED}:
        raise HTTPException(status_code=400, detail="Only draft or rejected budgets can be edited.")
    data = payload.model_dump(exclude_unset=True, by_alias=False)
    for attr in ("amount", "currency", "period", "justification"):
        if attr in data and data[attr] is not None:
            setattr(budget, attr, data[attr])
    db.add(budget)
    db.commit()
    return _serialize_budget(_load_budget(db, budget.id))


@router.post("/budgets/{budget_id}/submit")
def submit_budget(
    budget_id: str,
    db: Annotated[Session, Depends(get_db)],
    current_user: Annotated[User, Depends(require_permissions(Permission.PROJECTS_BUDGET_PROPOSE))],
) -> dict[str, Any]:
    budget = _load_budget(db, budget_id)
    project = _load_project(db, budget.project_id)
    pg.assert_can_act_on_project(db, current_user, project)
    pg.submit_budget(db, budget=budget, project=project, actor=current_user)
    db.commit()
    return _serialize_budget(_load_budget(db, budget.id))


@router.post("/budgets/{budget_id}/functional-decision")
def functional_decision(
    budget_id: str,
    payload: BudgetDecision,
    db: Annotated[Session, Depends(get_db)],
    current_user: Annotated[User, Depends(require_permissions(Permission.PROJECTS_BUDGET_APPROVE_FUNCTIONAL))],
) -> dict[str, Any]:
    budget = _load_budget(db, budget_id)
    project = _load_project(db, budget.project_id)
    # Identity gate: only the resolved CTO/COO (or admin/leadership) may act.
    if not pg.is_full_access(current_user) and budget.functional_approver_id != current_user.id:
        raise HTTPException(status_code=403, detail="Only the assigned functional approver (CTO/COO) can decide this budget.")
    approve = _decision_is_approve(payload)
    pg.apply_functional_decision(db, budget=budget, project=project, actor=current_user, approve=approve, comment=_clean(payload.comment))
    db.commit()
    return _serialize_budget(_load_budget(db, budget.id))


@router.post("/budgets/{budget_id}/leadership-decision")
def leadership_decision(
    budget_id: str,
    payload: BudgetDecision,
    db: Annotated[Session, Depends(get_db)],
    current_user: Annotated[User, Depends(require_permissions(Permission.PROJECTS_BUDGET_APPROVE_LEADERSHIP))],
) -> dict[str, Any]:
    budget = _load_budget(db, budget_id)
    project = _load_project(db, budget.project_id)
    approve = _decision_is_approve(payload)
    pg.apply_leadership_decision(db, budget=budget, project=project, actor=current_user, approve=approve, comment=_clean(payload.comment))
    db.commit()
    return _serialize_budget(_load_budget(db, budget.id))


def _decision_is_approve(payload: BudgetDecision) -> bool:
    action = (payload.action or "").strip().lower()
    if action in {"approve", "approved", "accept"}:
        return True
    if action in {"reject", "rejected", "decline"}:
        return False
    raise HTTPException(status_code=422, detail="action must be 'approve' or 'reject'.")


def _load_budget(db: Session, budget_id: str) -> ProjectBudget:
    budget = db.scalar(
        select(ProjectBudget)
        .options(
            selectinload(ProjectBudget.actions),
            selectinload(ProjectBudget.proposer),
            selectinload(ProjectBudget.functional_approver),
            selectinload(ProjectBudget.functional_decider),
            selectinload(ProjectBudget.leadership_decider),
        )
        .where(ProjectBudget.id == budget_id)
    )
    if budget is None:
        raise HTTPException(status_code=404, detail="Budget not found.")
    return budget


# --------------------------------------------------------------------------- #
# Analytics dashboard
# --------------------------------------------------------------------------- #
@router.get("/analytics/summary")
def analytics_summary(
    db: Annotated[Session, Depends(get_db)],
    current_user: Annotated[User, Depends(require_permissions(Permission.PROJECTS_READ))],
) -> dict[str, Any]:
    projects = list(db.scalars(_project_query().where(Project.is_archived.is_(False)).order_by(Project.internal_name)).unique())
    # Scope aggregates: Manager / PL see only their tagged projects.
    visible = pg.visible_project_ids(db, current_user)
    if visible is not None:
        projects = [p for p in projects if p.id in visible]

    total = len(projects)
    delivered = sum(1 for p in projects if p.rfp_status == "delivered" or p.delivery_status == "completed")
    active = sum(1 for p in projects if p.delivery_status != "completed")
    technical = sum(1 for p in projects if (p.project_type or "") == "technical")
    generalist = total - technical

    project_rows: list[dict[str, Any]] = []
    total_approved = 0.0
    total_consumed = 0.0
    tpm_portfolio: dict[str, dict[str, Any]] = defaultdict(lambda: {"projects": 0, "budget": 0.0})
    pl_portfolio: dict[str, dict[str, Any]] = defaultdict(lambda: {"projects": 0, "budget": 0.0})
    client_portfolio: dict[str, dict[str, Any]] = defaultdict(lambda: {"projects": 0, "budget": 0.0})

    # Per-project expense counts in two grouped queries.
    reimb_counts = dict(
        db.execute(
            select(ReimbursementRequest.project_id, func.count())
            .where(ReimbursementRequest.project_id.isnot(None))
            .group_by(ReimbursementRequest.project_id)
        ).all()
    )
    dinner_counts = dict(
        db.execute(
            select(DinnerRequest.project_id, func.count())
            .where(DinnerRequest.project_id.isnot(None))
            .group_by(DinnerRequest.project_id)
        ).all()
    )

    for p in projects:
        spend = pg.compute_project_spend(db, p.id)
        approved = p.approved_budget or 0.0
        total_approved += approved
        total_consumed += spend
        tpm_name = p.tpm.name if p.tpm else "Unassigned"
        tpm_portfolio[tpm_name]["projects"] += 1
        tpm_portfolio[tpm_name]["budget"] += approved
        for lead in (p.leads or []):
            if lead.role == "pl" and lead.user:
                pl_portfolio[lead.user.name]["projects"] += 1
                pl_portfolio[lead.user.name]["budget"] += approved
        client_key = p.client or p.external_name or "Internal"
        client_portfolio[client_key]["projects"] += 1
        client_portfolio[client_key]["budget"] += approved
        project_rows.append(
            {
                "id": p.id,
                "internalName": p.internal_name,
                "client": p.client,
                "projectType": p.project_type,
                "approvedBudget": approved,
                "consumedBudget": spend,
                "remainingBudget": round(approved - spend, 2),
                "memberCount": p.total_members or ((p.fte_count or 0) + (p.intern_count or 0)),
                "reimbursementCount": int(reimb_counts.get(p.id, 0)),
                "dinnerCount": int(dinner_counts.get(p.id, 0)),
            }
        )

    monthly = _monthly_expense_trend(db)

    def _portfolio_list(source: dict[str, dict[str, Any]]) -> list[dict[str, Any]]:
        return sorted(
            ({"name": k, **v} for k, v in source.items()),
            key=lambda r: r["budget"], reverse=True,
        )

    return {
        "totals": {
            "totalProjects": total,
            "activeProjects": active,
            "deliveredProjects": delivered,
            "totalApprovedBudget": round(total_approved, 2),
            "totalConsumedBudget": round(total_consumed, 2),
            "remainingBudget": round(total_approved - total_consumed, 2),
            "technical": technical,
            "generalist": generalist,
        },
        "projects": project_rows,
        "monthlyExpenseTrend": monthly,
        "tpmPortfolio": _portfolio_list(tpm_portfolio),
        "plPortfolio": _portfolio_list(pl_portfolio),
        "clientPortfolio": _portfolio_list(client_portfolio),
        "typeBreakdown": [
            {"name": "Technical", "value": technical},
            {"name": "Generalist", "value": generalist},
        ],
    }


def _monthly_expense_trend(db: Session) -> list[dict[str, Any]]:
    """Last 12 months of paid reimbursement + completed dinner spend."""
    buckets: dict[str, float] = defaultdict(float)
    reimbursements = db.execute(
        select(ReimbursementRequest.paid_at, ReimbursementRequest.expense_amount).where(
            ReimbursementRequest.status == "paid",
            ReimbursementRequest.project_id.isnot(None),
            ReimbursementRequest.paid_at.isnot(None),
        )
    ).all()
    for paid_at, amount in reimbursements:
        if paid_at:
            buckets[paid_at.strftime("%Y-%m")] += float(amount or 0.0)
    dinners = db.execute(
        select(DinnerRequest.completed_at, DinnerRequest.amount).where(
            DinnerRequest.status == "completed",
            DinnerRequest.project_id.isnot(None),
            DinnerRequest.amount.isnot(None),
            DinnerRequest.completed_at.isnot(None),
        )
    ).all()
    for completed_at, amount in dinners:
        if completed_at:
            buckets[completed_at.strftime("%Y-%m")] += float(amount or 0.0)
    return [{"month": month, "spend": round(buckets[month], 2)} for month in sorted(buckets)][-12:]


@router.get("/analytics/leadership")
def leadership_view(
    db: Annotated[Session, Depends(get_db)],
    current_user: Annotated[User, Depends(require_permissions(Permission.PROJECTS_READ))],
) -> dict[str, Any]:
    summary = analytics_summary(db, current_user)
    rows = summary["projects"]
    top_costing = sorted(rows, key=lambda r: r["consumedBudget"], reverse=True)[:10]
    profitability = sorted(
        (
            {
                "id": r["id"],
                "internalName": r["internalName"],
                "approvedBudget": r["approvedBudget"],
                "consumedBudget": r["consumedBudget"],
                "remainingBudget": r["remainingBudget"],
                "utilization": round((r["consumedBudget"] / r["approvedBudget"] * 100), 1) if r["approvedBudget"] else 0.0,
            }
            for r in rows
        ),
        key=lambda r: r["utilization"], reverse=True,
    )
    queue = db.scalars(
        select(ProjectBudget)
        .options(selectinload(ProjectBudget.proposer), selectinload(ProjectBudget.functional_approver))
        .where(ProjectBudget.status.in_([pg.STATUS_PENDING_FUNCTIONAL, pg.STATUS_PENDING_LEADERSHIP]))
        .order_by(ProjectBudget.submitted_at)
    )
    return {
        "totals": summary["totals"],
        "topCosting": top_costing,
        "profitability": profitability,
        "approvalQueue": [_serialize_budget(b) for b in queue],
    }


# --------------------------------------------------------------------------- #
# Escalations
# --------------------------------------------------------------------------- #
@router.post("/escalations/run")
def run_escalations(
    db: Annotated[Session, Depends(get_db)],
    current_user: Annotated[User, Depends(require_permissions(Permission.PROJECTS_ADMIN))],
) -> dict[str, Any]:
    result = pg.run_escalation_sweep(db)
    db.commit()
    return result


# --------------------------------------------------------------------------- #
# Excel / CSV import & export
# --------------------------------------------------------------------------- #
# Maps normalised sheet headers -> core project attribute.
_HEADER_TO_FIELD: dict[str, str] = {
    "internal": "internal_name",
    "internal_name": "internal_name",
    "external_client": "external_name",
    "external": "external_name",
    "client": "client",
    "rfp_production_delivered": "rfp_status",
    "appsheet_approval": "appsheet_approval",
    "trajectory_cost_approval": "trajectory_cost_approval",
    "platform": "platform",
    "techincal_generalist": "project_type",
    "technical_generalist": "project_type",
    "project_type": "project_type",
    "aht": "aht",
    "target_volume": "target_volume",
    "delivered": "delivered_volume",
    "date_of_delivery": "date_of_delivery",
    "tpm": "_tpm_name",
    "pl": "_pl_names",
    "ftes_resource_demand": "fte_demand",
    "no_of_ftes": "fte_count",
    "no_of_interns": "intern_count",
    "total_members": "total_members",
    "completed_ongoing": "delivery_status",
    "approved_budget": "approved_budget",
    "consumed_budget": "consumed_budget",
}


def _resolve_user_by_name(db: Session, name: str | None) -> User | None:
    if not name:
        return None
    cleaned = name.strip()
    if not cleaned:
        return None
    return db.scalar(select(User).where(func.lower(User.name) == cleaned.lower(), User.is_active.is_(True)))


@router.get("/export/file")
def export_projects(
    db: Annotated[Session, Depends(get_db)],
    current_user: Annotated[User, Depends(require_permissions(Permission.PROJECTS_READ))],
    fmt: Annotated[str, Query(alias="format")] = "csv",
) -> StreamingResponse:
    projects = list(db.scalars(_project_query().order_by(Project.internal_name)).unique())
    field_defs = list(db.scalars(select(ProjectFieldDef).where(ProjectFieldDef.is_active.is_(True)).order_by(ProjectFieldDef.order_index)))

    core_headers = [
        ("Internal", "internal_name"), ("External/Client", "external_name"), ("Client", "client"),
        ("Platform", "platform"), ("Technical / Generalist", "project_type"),
        ("RFP/Production/Delivered", "rfp_status"), ("Completed / Ongoing", "delivery_status"),
        ("Appsheet Approval", "appsheet_approval"), ("Trajectory Cost Approval", "trajectory_cost_approval"),
        ("AHT", "aht"), ("Target Volume", "target_volume"), ("Delivered", "delivered_volume"),
        ("Date of Delivery", "date_of_delivery"), ("TPM", "_tpm_name"), ("PL", "_pl_names"),
        ("FTEs Resource Demand", "fte_demand"), ("No of FTEs", "fte_count"), ("No. of Interns", "intern_count"),
        ("Total members", "total_members"), ("Approved Budget", "approved_budget"), ("Consumed Budget", "consumed_budget"),
    ]
    headers = [h for h, _ in core_headers] + [fd.label for fd in field_defs]

    def cell(project: Project, attr: str) -> Any:
        if attr == "_tpm_name":
            return project.tpm.name if project.tpm else ""
        if attr == "_pl_names":
            return ", ".join(lead.user.name for lead in (project.leads or []) if lead.role == "pl" and lead.user)
        if attr == "date_of_delivery":
            return project.date_of_delivery.isoformat() if project.date_of_delivery else ""
        if attr == "consumed_budget":
            return pg.compute_project_spend(db, project.id)
        value = getattr(project, attr, None)
        return "" if value is None else value

    rows = []
    for p in projects:
        row = [cell(p, attr) for _, attr in core_headers]
        row += [(p.custom_fields or {}).get(fd.key, "") for fd in field_defs]
        rows.append(row)

    stamp = datetime.now(UTC).date().isoformat()
    if fmt == "xlsx":
        try:
            from openpyxl import Workbook
        except ImportError:
            raise HTTPException(status_code=503, detail="Excel export unavailable (openpyxl not installed). Use format=csv.") from None
        wb = Workbook()
        ws = wb.active
        ws.title = "Projects"
        ws.append(headers)
        for row in rows:
            ws.append([str(c) if isinstance(c, (list, dict)) else c for c in row])
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return StreamingResponse(
            buffer,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f'attachment; filename="projects_{stamp}.xlsx"'},
        )

    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(headers)
    for row in rows:
        writer.writerow(["" if c is None else c for c in row])
    output.seek(0)
    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="text/csv",
        headers={"Content-Disposition": f'attachment; filename="projects_{stamp}.csv"'},
    )


@router.post("/bulk-upload")
def bulk_upload_projects(
    db: Annotated[Session, Depends(get_db)],
    current_user: Annotated[User, Depends(require_permissions(Permission.PROJECTS_CREATE))],
    file: Annotated[UploadFile, File()],
    auto_create_columns: Annotated[bool, Query(alias="autoCreateColumns")] = True,
) -> dict[str, Any]:
    rows = _read_tabular(file)
    if not rows:
        raise HTTPException(status_code=422, detail="No rows found in the uploaded file.")

    headers = list(rows[0].keys())
    field_defs = {fd.key: fd for fd in db.scalars(select(ProjectFieldDef))}
    label_to_key = {fd.label.strip().lower(): fd.key for fd in field_defs.values()}

    # Decide how each header maps: core field, known custom column, or new column.
    header_plan: dict[str, tuple[str, str]] = {}  # header -> (kind, target)
    max_order = db.scalar(select(func.coalesce(func.max(ProjectFieldDef.order_index), 0))) or 0
    for header in headers:
        norm = _slugify(header)
        if norm in _HEADER_TO_FIELD:
            header_plan[header] = ("core", _HEADER_TO_FIELD[norm])
            continue
        if norm in field_defs:
            header_plan[header] = ("custom", norm)
        elif header.strip().lower() in label_to_key:
            header_plan[header] = ("custom", label_to_key[header.strip().lower()])
        elif auto_create_columns and _clean(header):
            max_order += 1
            fd = ProjectFieldDef(
                id=generate_id(), key=norm, label=header.strip(), data_type="text",
                order_index=max_order, created_by=current_user.id, group="Imported",
            )
            db.add(fd)
            field_defs[norm] = fd
            header_plan[header] = ("custom", norm)
        else:
            header_plan[header] = ("ignore", "")

    created = updated = rejected = 0
    errors: list[dict[str, Any]] = []
    for index, raw in enumerate(rows, start=2):  # row 1 is the header
        core: dict[str, Any] = {}
        custom: dict[str, Any] = {}
        tpm_name = None
        pl_names: list[str] = []
        for header, value in raw.items():
            kind, target = header_plan.get(header, ("ignore", ""))
            if kind == "core":
                if target == "_tpm_name":
                    tpm_name = _clean(value)
                elif target == "_pl_names":
                    pl_names = [n.strip() for n in re.split(r"[,/;]", str(value or "")) if n.strip()]
                else:
                    core[target] = value
            elif kind == "custom":
                cleaned = _clean(value)
                if cleaned is not None:
                    custom[target] = cleaned

        internal_name = _clean(core.get("internal_name"))
        if not internal_name:
            rejected += 1
            errors.append({"row": index, "error": "Missing Internal project name."})
            continue

        project = db.scalar(select(Project).where(func.lower(Project.internal_name) == internal_name.lower()))
        is_new = project is None
        if is_new:
            project = Project(id=generate_id(), internal_name=internal_name, created_by=current_user.id)
            db.add(project)

        # Apply core attributes.
        project.internal_name = internal_name
        for attr in ("external_name", "client", "platform", "appsheet_approval", "trajectory_cost_approval"):
            if attr in core:
                setattr(project, attr, _clean(core[attr]))
        if "project_type" in core:
            project.project_type = _norm_enum(core["project_type"], PROJECT_TYPES, "technical")
        if "rfp_status" in core:
            project.rfp_status = _norm_enum(core["rfp_status"], RFP_STATUSES, "rfp")
        if "delivery_status" in core:
            project.delivery_status = _norm_enum(core["delivery_status"], DELIVERY_STATUSES, "ongoing")
        if "aht" in core:
            project.aht = _parse_float(core["aht"])
        for attr in ("target_volume", "delivered_volume", "fte_demand", "fte_count", "intern_count", "total_members"):
            if attr in core:
                setattr(project, attr, _parse_int(core[attr]))
        if "approved_budget" in core:
            project.approved_budget = _parse_float(core["approved_budget"])
        if "consumed_budget" in core:
            project.consumed_budget = _parse_float(core["consumed_budget"])
        if "date_of_delivery" in core:
            project.date_of_delivery = _parse_date(core["date_of_delivery"])
        if not project.total_members:
            project.total_members = (project.fte_count or 0) + (project.intern_count or 0) or None
        if custom:
            merged = dict(project.custom_fields or {})
            merged.update(custom)
            # keep raw names for display fallback when users can't be resolved
            if tpm_name:
                merged["_tpm_name"] = tpm_name
            if pl_names:
                merged["_pl_names"] = ", ".join(pl_names)
            project.custom_fields = merged
        elif tpm_name or pl_names:
            merged = dict(project.custom_fields or {})
            if tpm_name:
                merged["_tpm_name"] = tpm_name
            if pl_names:
                merged["_pl_names"] = ", ".join(pl_names)
            project.custom_fields = merged

        db.flush()

        # Resolve TPM/PL names to users where possible and (re)build leads.
        tpm_user = _resolve_user_by_name(db, tpm_name)
        if tpm_user:
            project.tpm_user_id = tpm_user.id
        pl_users = [u for u in (_resolve_user_by_name(db, n) for n in pl_names) if u]
        if tpm_user or pl_users:
            _replace_leads(db, project, tpm_user_id=project.tpm_user_id, lead_user_ids=[u.id for u in pl_users])

        if is_new:
            created += 1
        else:
            updated += 1

    db.commit()
    return {"total": len(rows), "created": created, "updated": updated, "rejected": rejected, "errors": errors[:50]}


def _read_tabular(file: UploadFile) -> list[dict[str, Any]]:
    """Parse an uploaded .csv or .xlsx into a list of header->value dicts."""
    raw = file.file.read()
    name = (file.filename or "").lower()
    if name.endswith(".xlsx") or name.endswith(".xlsm"):
        try:
            from openpyxl import load_workbook
        except ImportError:
            raise HTTPException(status_code=503, detail="Excel import unavailable (openpyxl not installed). Upload a CSV instead.") from None
        wb = load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
        ws = wb.active
        rows_iter = ws.iter_rows(values_only=True)
        try:
            header = [str(h).strip() if h is not None else "" for h in next(rows_iter)]
        except StopIteration:
            return []
        result = []
        for values in rows_iter:
            if values is None or all(v is None for v in values):
                continue
            row = {header[i]: values[i] if i < len(values) else None for i in range(len(header)) if header[i]}
            result.append(row)
        return result

    # CSV (default)
    text = raw.decode("utf-8-sig", errors="replace")
    reader = csv.DictReader(io.StringIO(text))
    return [dict(row) for row in reader if any((v or "").strip() for v in row.values())]


# --------------------------------------------------------------------------- #
# Public budget approve/reject link (dormant while email dispatch is off)
# --------------------------------------------------------------------------- #
def _html(message: str, *, ok: bool = True) -> HTMLResponse:
    color = "#16a34a" if ok else "#dc2626"
    return HTMLResponse(
        f"""<!doctype html><html><head><meta charset="utf-8"/>
        <meta name="viewport" content="width=device-width, initial-scale=1"/>
        <title>Budget Approval</title></head>
        <body style="font-family:system-ui,Arial,sans-serif;background:#0b0b14;color:#e5e7eb;
        display:flex;align-items:center;justify-content:center;height:100vh;margin:0">
        <div style="max-width:480px;padding:2rem;border:1px solid #2a2a3a;border-radius:16px;background:#14141f;text-align:center">
        <div style="font-size:2rem;color:{color};margin-bottom:.5rem">{'✓' if ok else '✕'}</div>
        <p style="font-size:1.05rem;line-height:1.5">{message}</p>
        </div></body></html>"""
    )


@public_router.get("/budget-approval", response_class=HTMLResponse)
def handle_budget_approval_link(
    token: str,
    db: Annotated[Session, Depends(get_db)],
) -> HTMLResponse:
    settings = get_settings()
    try:
        payload = decode_token(token, secret=settings.jwt_secret)
    except HTTPException:
        return _html("This approval link is invalid or has expired.", ok=False)
    if payload.get("type") != "project_budget_approval":
        return _html("This approval link is not valid.", ok=False)

    budget = db.get(ProjectBudget, payload.get("sub"))
    if budget is None:
        return _html("The budget for this link no longer exists.", ok=False)
    project = db.get(Project, budget.project_id)
    if project is None:
        return _html("The project for this link no longer exists.", ok=False)

    stage = payload.get("stage")
    request_id = payload.get("rid") or ""
    decision = payload.get("decision")
    stored_hash = budget.functional_token_hash if stage == "functional" else budget.leadership_token_hash
    expires_at = budget.functional_token_expires_at if stage == "functional" else budget.leadership_token_expires_at

    if not stored_hash or not verify_token_hash(request_id, stored_hash):
        return _html("This approval link has already been used or is no longer valid.", ok=False)
    if expires_at and expires_at < datetime.now(UTC):
        return _html("This approval link has expired.", ok=False)

    expected_status = pg.STATUS_PENDING_FUNCTIONAL if stage == "functional" else pg.STATUS_PENDING_LEADERSHIP
    if budget.status != expected_status:
        return _html("This budget has already been decided.", ok=False)

    approve = decision == "approve"
    if stage == "functional":
        pg.apply_functional_decision(db, budget=budget, project=project, actor=None, approve=approve, comment="Decided via email link.")
    else:
        pg.apply_leadership_decision(db, budget=budget, project=project, actor=None, approve=approve, comment="Decided via email link.")
    db.commit()
    verb = "approved" if approve else "rejected"
    return _html(f"Budget for '{project.internal_name}' has been {verb}. You can close this window.", ok=approve)
