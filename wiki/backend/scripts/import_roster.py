"""
One-shot import script for the Employee Birthday Roster.
Reads /tmp/dob.xlsx (or path from EXCEL_PATH env var) and upserts each row
into the local `employee_roster` collection keyed by `ecode`.

Run:  cd /app/backend && python -m scripts.import_roster
"""
import os
import sys
from datetime import datetime
from pathlib import Path

# Make backend imports available
sys.path.insert(0, str(Path(__file__).resolve().parents[1]))

import openpyxl
from dotenv import load_dotenv
from database import create_database

load_dotenv(Path(__file__).resolve().parents[1] / ".env")

EXCEL_PATH = os.environ.get("EXCEL_PATH", "/tmp/dob.xlsx")


def parse_dob(value):
    """Return YYYY-MM-DD string from datetime/date/str/number. Returns None if unparseable."""
    if value is None or value == "":
        return None
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d")
    if hasattr(value, "year") and hasattr(value, "month") and hasattr(value, "day"):
        return f"{value.year:04d}-{value.month:02d}-{value.day:02d}"
    if isinstance(value, str):
        # try common formats
        for fmt in ("%Y-%m-%d", "%d-%m-%Y", "%d/%m/%Y", "%m/%d/%Y", "%d %b %Y", "%d-%b-%Y"):
            try:
                return datetime.strptime(value.strip(), fmt).strftime("%Y-%m-%d")
            except ValueError:
                continue
    return None


def main():
    if not os.path.exists(EXCEL_PATH):
        print(f"ERROR: file not found at {EXCEL_PATH}")
        sys.exit(1)

    _client, db, backend_name = create_database(Path(__file__).resolve().parents[1])
    print(f"Using backend store: {backend_name}")

    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
    ws = wb[wb.sheetnames[0]]

    inserted = 0
    updated = 0
    skipped = 0
    failures = []

    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or all(v is None for v in row):
            continue
        sno, ecode, name, dob = (row + (None, None, None, None))[:4]
        ecode = str(ecode).strip() if ecode is not None else None
        name = str(name).strip() if name is not None else None
        dob_str = parse_dob(dob)
        if not ecode or not name:
            skipped += 1
            continue
        if not dob_str:
            failures.append({"ecode": ecode, "name": name, "raw_dob": str(dob)})
            skipped += 1
            continue
        doc = {
            "ecode": ecode,
            "name": name,
            "dob": dob_str,
            "department": "People Operations",
            "updated_at": datetime.utcnow().isoformat(),
        }
        result = db.employee_roster.update_one(
            {"ecode": ecode}, {"$set": doc}, upsert=True
        )
        if result.upserted_id is not None:
            inserted += 1
        elif result.modified_count > 0:
            updated += 1
        else:
            updated += 1  # already current

    # Ensure index for fast birthday lookup
    db.employee_roster.create_index("ecode", unique=True)
    db.employee_roster.create_index("dob")

    total = db.employee_roster.count_documents({})
    print(f"\nImport complete:")
    print(f"  inserted: {inserted}")
    print(f"  updated:  {updated}")
    print(f"  skipped:  {skipped}")
    print(f"  total in DB: {total}")
    if failures:
        print(f"  failed rows (first 5): {failures[:5]}")


if __name__ == "__main__":
    main()
